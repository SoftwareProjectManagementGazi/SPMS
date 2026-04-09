import io
import logging
from datetime import date as date_type, datetime
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from pathlib import Path

logger = logging.getLogger(__name__)

from app.api.dependencies import (
    get_current_user,
    get_project_repo,
    get_report_repo,
    _is_admin,
)
from app.application.dtos.report_dtos import (
    SummaryDTO,
    BurndownDTO,
    VelocityDTO,
    DistributionDTO,
    PerformanceDTO,
)
from app.application.use_cases.generate_reports import (
    GetSummaryUseCase,
    GetBurndownUseCase,
    GetVelocityUseCase,
    GetDistributionUseCase,
    GetPerformanceUseCase,
)
from app.domain.entities.user import User
from app.domain.repositories.project_repository import IProjectRepository
from app.domain.repositories.report_repository import IReportRepository

router = APIRouter()


def _parse_assignee_ids(assignee_ids: Optional[str]) -> Optional[List[int]]:
    """Parse comma-separated assignee IDs string into a list of ints.
    Returns None if input is None or empty string.
    """
    if not assignee_ids:
        return None
    try:
        ids = [int(i.strip()) for i in assignee_ids.split(",") if i.strip()]
        return ids if ids else None
    except ValueError:
        return None


@router.get("/summary", response_model=SummaryDTO)
async def get_summary(
    project_id: int,
    assignee_ids: Optional[str] = Query(None),
    date_from: Optional[date_type] = Query(None),
    date_to: Optional[date_type] = Query(None),
    current_user: User = Depends(get_current_user),
    project_repo: IProjectRepository = Depends(get_project_repo),
    report_repo: IReportRepository = Depends(get_report_repo),
):
    if not _is_admin(current_user):
        project = await project_repo.get_by_id_and_user(project_id, current_user.id)
        if project is None:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not a project member")

    parsed_ids = _parse_assignee_ids(assignee_ids)
    use_case = GetSummaryUseCase(report_repo)
    return await use_case.execute(project_id, parsed_ids, date_from, date_to)


@router.get("/burndown", response_model=BurndownDTO)
async def get_burndown(
    project_id: int,
    sprint_id: Optional[int] = Query(None),
    current_user: User = Depends(get_current_user),
    project_repo: IProjectRepository = Depends(get_project_repo),
    report_repo: IReportRepository = Depends(get_report_repo),
):
    if not _is_admin(current_user):
        project = await project_repo.get_by_id_and_user(project_id, current_user.id)
        if project is None:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not a project member")

    use_case = GetBurndownUseCase(report_repo)
    return await use_case.execute(project_id, sprint_id)


@router.get("/velocity", response_model=VelocityDTO)
async def get_velocity(
    project_id: int,
    date_from: Optional[date_type] = Query(None),
    date_to: Optional[date_type] = Query(None),
    current_user: User = Depends(get_current_user),
    project_repo: IProjectRepository = Depends(get_project_repo),
    report_repo: IReportRepository = Depends(get_report_repo),
):
    if not _is_admin(current_user):
        project = await project_repo.get_by_id_and_user(project_id, current_user.id)
        if project is None:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not a project member")

    use_case = GetVelocityUseCase(report_repo)
    return await use_case.execute(project_id, date_from, date_to)


@router.get("/distribution", response_model=DistributionDTO)
async def get_distribution(
    project_id: int,
    group_by: str = Query("status"),
    assignee_ids: Optional[str] = Query(None),
    date_from: Optional[date_type] = Query(None),
    date_to: Optional[date_type] = Query(None),
    current_user: User = Depends(get_current_user),
    project_repo: IProjectRepository = Depends(get_project_repo),
    report_repo: IReportRepository = Depends(get_report_repo),
):
    if not _is_admin(current_user):
        project = await project_repo.get_by_id_and_user(project_id, current_user.id)
        if project is None:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not a project member")

    parsed_ids = _parse_assignee_ids(assignee_ids)
    use_case = GetDistributionUseCase(report_repo)
    return await use_case.execute(project_id, group_by, parsed_ids, date_from, date_to)


@router.get("/performance", response_model=PerformanceDTO)
async def get_performance(
    project_id: int,
    assignee_ids: Optional[str] = Query(None),
    date_from: Optional[date_type] = Query(None),
    date_to: Optional[date_type] = Query(None),
    current_user: User = Depends(get_current_user),
    project_repo: IProjectRepository = Depends(get_project_repo),
    report_repo: IReportRepository = Depends(get_report_repo),
):
    if not _is_admin(current_user):
        project = await project_repo.get_by_id_and_user(project_id, current_user.id)
        if project is None:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not a project member")

    parsed_ids = _parse_assignee_ids(assignee_ids)
    use_case = GetPerformanceUseCase(report_repo)
    return await use_case.execute(project_id, parsed_ids, date_from, date_to)


@router.get("/export/pdf")
async def export_pdf(
    project_id: int,
    assignee_ids: Optional[str] = Query(None),
    date_from: Optional[date_type] = Query(None),
    date_to: Optional[date_type] = Query(None),
    current_user: User = Depends(get_current_user),
    project_repo: IProjectRepository = Depends(get_project_repo),
    report_repo: IReportRepository = Depends(get_report_repo),
):
    if not _is_admin(current_user):
        project = await project_repo.get_by_id_and_user(project_id, current_user.id)
        if project is None:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not a project member")

    try:
        project_obj = await project_repo.get_by_id(project_id)
        if not project_obj:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")

        parsed_ids = _parse_assignee_ids(assignee_ids)
        tasks = await report_repo.get_tasks_for_export(project_id, parsed_ids, date_from, date_to)

        filter_parts = []
        if parsed_ids:
            filter_parts.append(f"Atanan: {', '.join(str(i) for i in parsed_ids)}")
        if date_from:
            filter_parts.append(f"Baslangic: {date_from}")
        if date_to:
            filter_parts.append(f"Bitis: {date_to}")
        filter_summary = " | ".join(filter_parts) if filter_parts else "Tum veriler"
        generated_at = datetime.utcnow().strftime("%Y-%m-%d %H:%M")

        from fpdf import FPDF

        UNICODE_FONT = "/Library/Fonts/Arial Unicode.ttf"
        pdf = FPDF(orientation="L", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=10)
        pdf.add_page()

        if Path(UNICODE_FONT).exists():
            pdf.add_font("uni", fname=UNICODE_FONT)
            fn = "uni"
        else:
            fn = "Helvetica"

        # Title
        pdf.set_font(fn, size=14)
        pdf.cell(0, 10, f"SPMS Raporu - {project_obj.name}", new_x="LMARGIN", new_y="NEXT")

        # Subtitle
        pdf.set_font(fn, size=9)
        pdf.cell(0, 6, f"Filtre: {filter_summary}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 6, f"Olusturulma: {generated_at} UTC", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

        # Column widths (landscape A4 = 287 mm usable)
        col_widths = [22, 75, 28, 38, 22, 34, 14, 28, 26]
        headers = ["Kod", "Baslik", "Durum", "Atanan", "Oncelik", "Sprint", "Puan", "Olusturulma", "Bitis"]

        # Header row
        pdf.set_fill_color(79, 70, 229)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font(fn, size=9)
        for w, h in zip(col_widths, headers):
            pdf.cell(w, 8, h, border=1, fill=True)
        pdf.ln()

        # Data rows
        pdf.set_text_color(0, 0, 0)
        pdf.set_font(fn, size=8)
        alt = False
        for task in tasks:
            pdf.set_fill_color(248, 248, 248) if alt else pdf.set_fill_color(255, 255, 255)
            row_vals = [
                task.task_key or "",
                (task.title or "")[:55],
                task.status or "",
                task.assignee or "",
                task.priority or "",
                task.sprint or "",
                str(task.points) if task.points is not None else "",
                task.created_at.strftime("%Y-%m-%d") if task.created_at else "",
                task.due_date.strftime("%Y-%m-%d") if task.due_date else "",
            ]
            for val, w in zip(row_vals, col_widths):
                pdf.cell(w, 7, str(val), border=1, fill=True)
            pdf.ln()
            alt = not alt

        pdf_bytes = bytes(pdf.output())
        project_key = getattr(project_obj, 'key', 'UNKNOWN')
        filename = f"SPMS_Report_{project_key}_{date_type.today()}.pdf"
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("PDF export failed")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"PDF olusturulamadi: {type(exc).__name__}: {exc}",
        )


@router.get("/export/excel")
async def export_excel(
    project_id: int,
    assignee_ids: Optional[str] = Query(None),
    date_from: Optional[date_type] = Query(None),
    date_to: Optional[date_type] = Query(None),
    current_user: User = Depends(get_current_user),
    project_repo: IProjectRepository = Depends(get_project_repo),
    report_repo: IReportRepository = Depends(get_report_repo),
):
    # Same membership check as PDF
    if not _is_admin(current_user):
        project = await project_repo.get_by_id_and_user(project_id, current_user.id)
        if project is None:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not a project member")

    try:
        project_obj = await project_repo.get_by_id(project_id)
        if not project_obj:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Project not found")

        parsed_ids = _parse_assignee_ids(assignee_ids)
        tasks = await report_repo.get_tasks_for_export(project_id, parsed_ids, date_from, date_to)

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SPMS Raporu"

        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        header_alignment = Alignment(horizontal="left", vertical="center")

        headers = ["Gorev Kodu", "Baslik", "Durum", "Atanan", "Oncelik",
                   "Sprint", "Puan", "Olusturulma", "Bitis Tarihi", "Guncelleme", "Raporlayan"]
        col_widths = [14, 40, 16, 24, 14, 20, 8, 18, 18, 18, 24]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 20

        for task in tasks:
            ws.append([
                task.task_key or "",
                task.title,
                task.status or "",
                task.assignee or "",
                task.priority or "",
                task.sprint or "",
                task.points,
                task.created_at.strftime("%Y-%m-%d %H:%M") if task.created_at else "",
                task.due_date.strftime("%Y-%m-%d") if task.due_date else "",
                task.updated_at.strftime("%Y-%m-%d %H:%M") if task.updated_at else "",
                task.reporter or "",
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        project_key = getattr(project_obj, 'key', 'UNKNOWN')
        filename = f"SPMS_Report_{project_key}_{date_type.today()}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Excel export failed")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Excel olusturulamadi: {type(exc).__name__}: {exc}",
        )
